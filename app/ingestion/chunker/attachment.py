"""첨부 1차 분할 + 첨부 청킹 엔트리 (feature4: pdf / docx / xlsx / csv).

--------------------------------------------------
작성자 : 최태성
작성목적 : 다운로드된 첨부 파일을 attachment_type별 전략으로 청크로 분할한다
          (chunking-strategy.md §5). pdf는 PyMuPDF(fitz)로 폰트 크기·굵기·짧은 행
          휴리스틱을 적용해 섹션을 잡고(미검출 시 슬라이딩 윈도우, 추출 실패 시
          pdfplumber 폴백), docx는 python-docx로 Heading 1/2/3 계층을 1차 분할 단위로
          삼고(표는 markdown 변환), xlsx는 openpyxl로 시트 단위 → 시트 내 N행 그룹으로
          분할해 각 행을 컬럼명과 함께 자연어로 직렬화한다. csv는 단일 시트로 보고
          xlsx 행 직렬화 자산을 재사용한다.
작성일 : 2026-05-15
변경사항 내역 (날짜, 변경목적, 변경내용 순)
  - 2026-05-15, 최초 작성, feature4-A — docx/xlsx 분할기 + chunk_attachment
  - 2026-05-17, 코드 리뷰 후속(P1-3·P2) — attachment.local_path 우선 사용(ADR-2026-001),
    xlsx 단일 행/축소 한계 그룹 oversize 슬라이딩 윈도우 분할, _looks_like_header가
    raw value 기반으로 datetime을 헤더로 오인하지 않도록 보강, doc_type enum 그대로 전달
  - 2026-05-22, feature4-B(csv) — csv 분할기 추가. 단일 시트로 보고 _resolve_header/
    _group_sheet_rows 재사용, 인코딩 자동감지(utf-8-sig/cp949 fallback, 의존성 없음),
    수치 문자열을 비헤더로 보도록 _looks_like_header 보강(xlsx 무회귀).
  - 2026-05-22, feature4-B(pdf) — pdf 분할기 추가. fitz로 폰트 휴리스틱 섹션 분할
    (section_header=p.<N>: <제목>), 헤딩 미검출 시 단일 draft→800토큰 슬라이딩 윈도우,
    fitz 추출 0건 시 pdfplumber 폴백, 암호화 PDF는 ATTACH_ENCRYPTED ValueError.
    feature4 (pdf/docx/xlsx/csv) 전부 완료.
--------------------------------------------------
[호환성]
  - Python 3.11.x, python-docx 1.1+, openpyxl 3.1+, pymupdf 1.24+, pdfplumber 0.11+.
    csv는 표준 라이브러리만 사용. pdfplumber는 폴백 경로에서만 지연 import한다.
--------------------------------------------------
"""

import csv
import io
import re
from collections import Counter
from collections.abc import Iterator
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import fitz
import openpyxl
from docx import Document as load_docx
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph

if TYPE_CHECKING:
    from docx.document import Document as DocxDocument

from app.ingestion.chunker.base import (
    MAX_TOKENS,
    ChunkDraft,
    apply_size_rules,
    split_oversized,
)
from app.ingestion.chunker.tokenizer import count_tokens
from app.schemas.chunk import Chunk, ChunkMetadata, make_chunk_id
from app.schemas.enums import AttachmentType, ExtractedFormat, SourceType
from app.schemas.page_object import Attachment, PageObject

# Heading 1/2/3만 docx 섹션 경계로 본다 (chunking-strategy.md §5: "Heading 1/2/3 → 단락 fallback").
_DOCX_HEADING_STYLE = re.compile(r"^Heading [123]$")

# xlsx 행 그룹 크기 — 직렬화 결과가 800토큰 초과 시 다음 단계로 축소 (chunking-strategy.md §5).
_XLSX_GROUP_SIZES = (50, 25, 10)

# csv 디코딩 인코딩 후보 — 앞에서부터 시도한다 (PoC 인코딩 자동감지, 의존성 없음).
# utf-8-sig는 Excel이 저장한 BOM을 제거하고, cp949는 국내 CSV에 흔하다.
_CSV_ENCODINGS = ("utf-8-sig", "cp949", "utf-8", "latin-1")

# 셀 문자열이 순수 수치로 보이는지 — 헤더 판정에서 수치 문자열을 데이터로 본다 (csv 보강).
_NUMERIC_RE = re.compile(r"^-?\d+(?:\.\d+)?$")

# attachment_type 판별용 mime 힌트 (PoC) — 실제 분류는 첨부 분석기 [Pipeline]=feature6 책임.
_MIME_HINTS = (
    ("pdf", AttachmentType.PDF),
    ("wordprocessingml", AttachmentType.DOCX),
    ("spreadsheetml", AttachmentType.XLSX),
    ("csv", AttachmentType.CSV),
)
_EXTENSION_TYPES = {
    ".pdf": AttachmentType.PDF,
    ".docx": AttachmentType.DOCX,
    ".xlsx": AttachmentType.XLSX,
    ".csv": AttachmentType.CSV,
}
_EXTRACTED_FORMAT_BY_TYPE = {
    AttachmentType.PDF: ExtractedFormat.RAW_TEXT,
    AttachmentType.DOCX: ExtractedFormat.RAW_TEXT,
    AttachmentType.XLSX: ExtractedFormat.SHEET_SERIALIZED,
    AttachmentType.CSV: ExtractedFormat.SHEET_SERIALIZED,
}

# pdf 헤딩 휴리스틱 임계 — 본문 폰트 대비 비율·짧은 행 기준 (chunking-strategy.md §5).
_PDF_HEADING_SIZE_RATIO = 1.15  # 본문보다 충분히 큰 폰트
_PDF_BOLD_SIZE_RATIO = 1.05  # 볼드는 약간만 커도 헤딩 후보
_PDF_HEADING_MAX_CHARS = 80  # 헤딩으로 볼 짧은 행 길이 상한
_PDF_HEADING_MAX_WORDS = 12
_PDF_BOLD_FLAG = 16  # PyMuPDF span flags의 bold 비트(2**4)


def infer_attachment_type(attachment: Attachment) -> AttachmentType:
    """mime_type·확장자로 attachment_type을 추정한다 — PoC 휴리스틱.

    실제 attachment_type은 첨부 파일 분석기 [Pipeline](feature6)가 결정한다. 본 함수는
    feature4 단독 테스트·데모를 위한 임시 추정기다.

    Args:
        attachment: 부모 페이지에 부속된 첨부 객체.

    Returns:
        추정된 AttachmentType.

    Raises:
        ValueError: mime·확장자 어느 쪽으로도 유형을 판별할 수 없을 때.
    """
    mime = attachment.mime_type.lower()
    for hint, attachment_type in _MIME_HINTS:
        if hint in mime:
            return attachment_type
    extension = Path(attachment.filename).suffix.lower()
    if extension in _EXTENSION_TYPES:
        return _EXTENSION_TYPES[extension]
    raise ValueError(
        f"첨부 유형을 판별할 수 없습니다 "
        f"(mime={attachment.mime_type}, filename={attachment.filename})"
    )


def _coerce_attachment_type(attachment_type: AttachmentType | str) -> AttachmentType:
    """attachment_type을 AttachmentType으로 변환한다. 미인식 값은 ValueError."""
    if isinstance(attachment_type, AttachmentType):
        return attachment_type
    try:
        return AttachmentType(attachment_type)
    except ValueError as exc:
        raise ValueError(f"알 수 없는 attachment_type: {attachment_type!r}") from exc


def _resolve_attachment_path(attachment: Attachment) -> str:
    """청커가 파일을 직접 열기 위한 경로를 결정한다 (ADR-2026-001).

    ``local_path``가 채워져 있으면 그것을 사용하고, 없으면 ``download_url``을 fallback으로
    사용한다(file:// scheme이면 그 경로를 가리킨다고 가정). 운영 어댑터가 도착하면 다운로드
    헬퍼가 ``local_path``를 채워주는 것이 정공법이다.
    """
    if attachment.local_path:
        return attachment.local_path
    url = attachment.download_url
    if url.startswith("file://"):
        from urllib.parse import urlparse

        return urlparse(url).path
    return url


# --- docx 1차 분할 ---


def _iter_block_items(document: "DocxDocument") -> Iterator[Paragraph | Table]:
    """docx 본문(body)의 문단·표를 문서에 나타난 순서대로 순회한다.

    python-docx의 `document.paragraphs` / `document.tables`는 순서 정보를 잃으므로,
    표가 어느 섹션에 속하는지 보존하려면 body XML 자식을 직접 순회해야 한다.
    """
    body = document.element.body
    for child in body.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, document)
        elif child.tag == qn("w:tbl"):
            yield Table(child, document)


def _is_docx_heading(paragraph: Paragraph) -> bool:
    """문단 스타일이 Heading 1/2/3이면 True (섹션 경계)."""
    style = paragraph.style
    return (
        style is not None and style.name is not None and bool(_DOCX_HEADING_STYLE.match(style.name))
    )


def _docx_table_to_markdown(table: Table) -> str:
    """docx 표를 markdown 표 텍스트로 변환한다. 셀 내 줄바꿈은 공백으로 치환한다."""
    rows = list(table.rows)
    if not rows:
        return ""
    lines: list[str] = []
    for row_index, row in enumerate(rows):
        cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
        lines.append("| " + " | ".join(cells) + " |")
        if row_index == 0:
            lines.append("| " + " | ".join("---" for _ in cells) + " |")
    return "\n".join(lines)


def _extract_docx_sections(path: str) -> tuple[list[str], list[tuple[str, list[str]]]]:
    """docx 본문을 (preamble 블록, [(heading, [본문 블록])]) 구조로 추출한다.

    Heading 1/2/3 이전의 표지 문단은 preamble로 모으고, 이후 문단·표는 직전 헤딩
    섹션에 누적한다. 표는 markdown으로 변환해 텍스트 블록으로 취급한다.
    """
    document = load_docx(path)
    preamble: list[str] = []
    sections: list[tuple[str, list[str]]] = []
    for block in _iter_block_items(document):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            if _is_docx_heading(block):
                sections.append((text, []))
            elif sections:
                sections[-1][1].append(text)
            else:
                preamble.append(text)
        else:
            markdown = _docx_table_to_markdown(block)
            if not markdown:
                continue
            if sections:
                sections[-1][1].append(markdown)
            else:
                preamble.append(markdown)
    return preamble, sections


def _chunk_docx(attachment: Attachment) -> list[ChunkDraft]:
    """docx 첨부를 Heading 1/2/3 섹션 단위로 1차 분할한다.

    헤딩이 하나도 없으면 전체를 단일 draft로 폴백하고 section_header에 파일명을 쓴다.
    첨부 섹션은 원자성이 없으므로 is_atomic=False — 크기 규칙은 chunk_attachment가 적용한다.
    """
    preamble, sections = _extract_docx_sections(_resolve_attachment_path(attachment))
    if not sections:
        text = "\n".join(preamble).strip()
        if not text:
            return []
        return [ChunkDraft(text=text, section_header=attachment.filename, is_atomic=False)]
    drafts: list[ChunkDraft] = []
    for index, (heading, body_blocks) in enumerate(sections):
        body = "\n".join(body_blocks).strip()
        text = f"{heading}\n{body}".strip() if body else heading
        # 표지 문단(preamble)은 첫 섹션 도입부에 부착한다 (맥락 동봉 — chunking-strategy.md §1).
        if index == 0 and preamble:
            text = "\n".join(preamble).strip() + "\n\n" + text
        drafts.append(ChunkDraft(text=text, section_header=heading, is_atomic=False))
    return drafts


# --- xlsx 1차 분할 ---


def _cell_to_str(value: object) -> str:
    """셀 값을 직렬화용 문자열로 변환한다. 빈 셀은 빈 문자열."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:g}"
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _looks_like_header(row: list[object]) -> bool:
    """첫 행이 헤더로 보이는지 판단한다 — 비어 있지 않은 셀이 모두 비수치·비-datetime 텍스트면 헤더.

    raw 셀 값(`int`/`float`/`datetime`)을 그대로 검사한다. ``_cell_to_str``이 datetime을
    isoformat 문자열로 미리 변환하면 datetime 셀이 텍스트로 보이는 false-positive가
    발생하므로, raw value 기반 판정으로 보강한다 (P2 보완, 2026-05-17).

    csv는 모든 셀이 문자열이라 raw 타입 판정만으로는 수치 행을 데이터로 구분할 수 없다.
    수치로 보이는 문자열(``_NUMERIC_RE``)도 비헤더로 본다 (feature4-B csv 보강).
    xlsx 헤더는 통상 설명 텍스트라 이 보강은 무회귀다.
    """
    non_empty = [cell for cell in row if cell is not None and _cell_to_str(cell) != ""]
    if not non_empty:
        return False
    for cell in non_empty:
        if isinstance(cell, (int, float, datetime)):
            return False
        if _NUMERIC_RE.match(_cell_to_str(cell)):
            return False
    return True


def _resolve_header(rows: list[list[object]]) -> tuple[list[str], list[list[object]]]:
    """시트 행 목록에서 컬럼명 헤더와 데이터 행을 분리한다.

    첫 행이 헤더로 보이면 헤더로 채택하고, 아니면 헤더 누락으로 보아 col_1, col_2...
    를 부여한다 (chunking-strategy.md §8 ATTACH_NO_HEADER).
    """
    width = max(len(row) for row in rows)
    if _looks_like_header(rows[0]):
        first = rows[0]
        header = [
            _cell_to_str(first[index]) if index < len(first) else "" for index in range(width)
        ]
        header = [name or f"col_{index + 1}" for index, name in enumerate(header)]
        return header, rows[1:]
    return [f"col_{index + 1}" for index in range(width)], rows


def _serialize_row(sheet_name: str, header: list[str], row: list[object]) -> str:
    """한 행을 `[<시트명>] <컬럼>: <값> | ...` 형식으로 직렬화한다. 빈 셀은 생략한다."""
    cells: list[str] = []
    for index, value in enumerate(row):
        text = _cell_to_str(value)
        if not text:
            continue
        column = header[index] if index < len(header) else f"col_{index + 1}"
        cells.append(f"{column}: {text}")
    return f"[{sheet_name}] " + " | ".join(cells)


def _serialize_rows(sheet_name: str, header: list[str], rows: list[list[object]]) -> str:
    """행 묶음을 줄바꿈으로 이어 직렬화한다 — 컬럼명은 매 행에 반복 부착된다."""
    return "\n".join(_serialize_row(sheet_name, header, row) for row in rows)


def _next_smaller_group_size(current: int) -> int | None:
    """현재 행 수보다 작은 다음 그룹 크기를 반환한다. 더 작은 단계가 없으면 None."""
    for size in _XLSX_GROUP_SIZES:
        if size < current:
            return size
    return None


def _group_sheet_rows(
    sheet_name: str,
    header: list[str],
    data_rows: list[list[object]],
) -> list[ChunkDraft]:
    """시트 데이터 행을 N행 그룹으로 묶어 ChunkDraft 목록을 만든다.

    기본 50행 그룹을 만들되, 직렬화 결과가 800토큰을 초과하면 25→10행으로 축소
    재분할한다. 10행 그룹도 초과하면 더 줄일 단계가 없으므로 단일 행 슬라이딩 윈도우
    분할(800토큰/100토큰 오버랩)을 적용해 임베딩 모델 입력 한계를 넘지 않게 한다
    (chunking-strategy.md §5, P2 보완 2026-05-17).
    """
    drafts: list[ChunkDraft] = []

    def emit_single_row(row: list[object], row_index: int) -> None:
        """단일 행도 800토큰을 넘으면 텍스트를 슬라이딩 윈도우로 추가 분할한다."""
        text = _serialize_rows(sheet_name, header, [row])
        if not text.strip():
            return
        section_header = f"[{sheet_name}] 행 {row_index + 1}"
        windows = split_oversized(text)
        if len(windows) == 1:
            drafts.append(ChunkDraft(text=text, section_header=section_header, is_atomic=False))
            return
        for part_index, window in enumerate(windows, start=1):
            part_header = f"{section_header} (part {part_index}/{len(windows)})"
            drafts.append(ChunkDraft(text=window, section_header=part_header, is_atomic=False))

    def emit(rows: list[list[object]], start_index: int) -> None:
        text = _serialize_rows(sheet_name, header, rows)
        if not text.strip():
            return
        smaller = _next_smaller_group_size(len(rows))
        if count_tokens(text) > MAX_TOKENS:
            if smaller is not None:
                for offset in range(0, len(rows), smaller):
                    emit(rows[offset : offset + smaller], start_index + offset)
                return
            # 더 줄일 그룹 단계가 없으면 행 단위로 분해 (단일 행도 oversize면 텍스트 슬라이딩).
            if len(rows) > 1:
                for row_offset, row in enumerate(rows):
                    emit_single_row(row, start_index + row_offset)
                return
            emit_single_row(rows[0], start_index)
            return
        section_header = f"[{sheet_name}] 행 {start_index + 1}~{start_index + len(rows)}"
        drafts.append(ChunkDraft(text=text, section_header=section_header, is_atomic=False))

    for offset in range(0, len(data_rows), _XLSX_GROUP_SIZES[0]):
        emit(data_rows[offset : offset + _XLSX_GROUP_SIZES[0]], offset)
    return drafts


def _chunk_xlsx(attachment: Attachment) -> list[ChunkDraft]:
    """xlsx 첨부를 시트 단위 → 시트 내 N행 그룹 단위로 1차 분할한다.

    빈 행은 제외하고, 각 행을 컬럼명과 함께 자연어로 직렬화한다. 행 그룹 분할 자체가
    xlsx의 크기 처리이므로 chunk_attachment는 별도 크기 규칙을 적용하지 않는다.
    """
    workbook = openpyxl.load_workbook(_resolve_attachment_path(attachment), data_only=True)
    drafts: list[ChunkDraft] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [
                list(row)
                for row in worksheet.iter_rows(values_only=True)
                if any(_cell_to_str(cell) != "" for cell in row)
            ]
            if not rows:
                continue
            header, data_rows = _resolve_header(rows)
            if not data_rows:
                continue
            drafts.extend(_group_sheet_rows(sheet_name, header, data_rows))
    finally:
        workbook.close()
    return drafts


# --- csv 1차 분할 ---


def _read_csv_rows(path: str) -> list[list[object]]:
    """csv 파일을 행 목록으로 읽는다 — 인코딩 자동감지(PoC) + 완전 빈 행 제외.

    ``_CSV_ENCODINGS`` 순으로 디코딩을 시도하고(utf-8-sig가 Excel BOM 제거), 모두
    실패하면 utf-8 + replace로 폴백한다. 셀 값은 좌우 공백을 제거해 문자열로 보존하며,
    수치/날짜 형태로 변환하지 않는다(원본 충실 — ID·선행 0 손실 방지). 반환 타입은
    xlsx 직렬화 헬퍼(`_resolve_header`/`_group_sheet_rows`)와 정합하도록 list[list[object]]다.
    """
    raw = Path(path).read_bytes()
    text: str | None = None
    for encoding in _CSV_ENCODINGS:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    rows: list[list[object]] = [
        [cell.strip() for cell in row] for row in csv.reader(io.StringIO(text))
    ]
    return [row for row in rows if any(cell != "" for cell in row)]


def _chunk_csv(attachment: Attachment) -> list[ChunkDraft]:
    """csv 첨부를 단일 시트로 보고 N행 그룹 단위로 1차 분할한다.

    xlsx의 행 직렬화 자산(``_resolve_header``/``_group_sheet_rows``)을 그대로 재사용한다.
    시트명이 없으므로 파일명 stem을 시트명으로 쓴다. xlsx와 마찬가지로 행 그룹 분할이
    크기 처리를 겸하므로 chunk_attachment는 별도 크기 규칙을 적용하지 않는다.
    """
    rows = _read_csv_rows(_resolve_attachment_path(attachment))
    if not rows:
        return []
    sheet_name = Path(attachment.filename).stem or attachment.filename
    header, data_rows = _resolve_header(rows)
    if not data_rows:
        return []
    return _group_sheet_rows(sheet_name, header, data_rows)


# --- pdf 1차 분할 ---


def _pdf_line_records(
    document: "fitz.Document",
) -> tuple[list[tuple[int, str, float, bool]], float]:
    """PDF 라인을 (page_no, text, max_size, bold) 레코드로 추출하고 본문 폰트 크기를 추정한다.

    본문 폰트 크기는 글자 수로 가중한 최빈 span 크기로 본다(가장 많은 글자를 차지하는 크기).
    bold는 span flags의 bold 비트 또는 폰트명에 'bold' 포함으로 판정한다.
    """
    records: list[tuple[int, str, float, bool]] = []
    size_chars: Counter[float] = Counter()
    for page_no, page in enumerate(document, start=1):
        for block in page.get_text("dict")["blocks"]:
            for line in block.get("lines", []):
                spans = line["spans"]
                text = "".join(span["text"] for span in spans).strip()
                if not text:
                    continue
                max_size = round(max(span["size"] for span in spans), 1)
                bold = any(
                    bool(int(span["flags"]) & _PDF_BOLD_FLAG) or "bold" in span["font"].lower()
                    for span in spans
                )
                records.append((page_no, text, max_size, bold))
                size_chars[round(min(span["size"] for span in spans), 1)] += len(text)
    body_size = size_chars.most_common(1)[0][0] if size_chars else 0.0
    return records, body_size


def _is_pdf_heading(text: str, size: float, bold: bool, body_size: float) -> bool:
    """라인이 섹션 헤딩으로 보이는지 판단한다 — 짧은 행 + 큰 폰트(또는 볼드)."""
    if body_size <= 0:
        return False
    if len(text) > _PDF_HEADING_MAX_CHARS or len(text.split()) > _PDF_HEADING_MAX_WORDS:
        return False
    if size >= body_size * _PDF_HEADING_SIZE_RATIO:
        return True
    return bool(bold and size >= body_size * _PDF_BOLD_SIZE_RATIO)


def _extract_pdf_sections(path: str) -> tuple[list[str], list[tuple[str, list[str]]]]:
    """PDF 본문을 (preamble 블록, [(section_header, [본문 블록])]) 구조로 추출한다.

    헤딩 검출 휴리스틱(폰트 크기·굵기·짧은 행)으로 섹션 경계를 잡고, section_header는
    ``p.<페이지>: <제목>`` 형식으로 만든다(chunking-strategy.md §5). 첫 헤딩 이전 라인은
    preamble로 모은다. 텍스트가 전혀 없으면 ([], [])를 돌려 호출자가 pdfplumber로 폴백한다.

    Raises:
        ValueError: 암호화되어 텍스트를 추출할 수 없는 PDF (ATTACH_ENCRYPTED).
    """
    document = fitz.open(path)
    try:
        if document.needs_pass:
            raise ValueError(f"ATTACH_ENCRYPTED: 암호화된 PDF는 처리할 수 없습니다 ({path})")
        records, body_size = _pdf_line_records(document)
    finally:
        document.close()
    preamble: list[str] = []
    sections: list[tuple[str, list[str]]] = []
    for page_no, text, size, bold in records:
        if _is_pdf_heading(text, size, bold, body_size):
            sections.append((f"p.{page_no}: {text}", []))
        elif sections:
            sections[-1][1].append(text)
        else:
            preamble.append(text)
    return preamble, sections


def _extract_pdf_plain_text(path: str) -> str:
    """pdfplumber로 PDF 전체 텍스트를 추출한다 — fitz가 텍스트를 못 뽑을 때의 폴백.

    pdfplumber는 무거운 폴백 경로이므로 지연 import한다(복잡 레이아웃 대비).
    """
    import pdfplumber

    parts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                parts.append(text)
    return "\n".join(parts)


def _chunk_pdf(attachment: Attachment) -> list[ChunkDraft]:
    """PDF 첨부를 폰트 휴리스틱 섹션 단위로 1차 분할한다.

    fitz로 헤딩을 검출해 섹션을 만들고(section_header=``p.<N>: <제목>``), 헤딩이 하나도
    없으면 전체를 단일 draft로 폴백한다(chunk_attachment의 2차 크기 규칙이 800토큰 슬라이딩
    윈도우로 재분할). fitz가 텍스트를 전혀 못 뽑으면 pdfplumber 폴백으로 평문을 추출한다.
    PDF 섹션은 원자성이 없으므로 is_atomic=False.
    """
    path = _resolve_attachment_path(attachment)
    preamble, sections = _extract_pdf_sections(path)
    if not preamble and not sections:
        # fitz 추출 0건 → pdfplumber 평문 폴백 (이미지/복잡 레이아웃 대비).
        text = _extract_pdf_plain_text(path).strip()
        if not text:
            return []
        return [ChunkDraft(text=text, section_header=attachment.filename, is_atomic=False)]
    if not sections:
        text = "\n".join(preamble).strip()
        if not text:
            return []
        return [ChunkDraft(text=text, section_header=attachment.filename, is_atomic=False)]
    drafts: list[ChunkDraft] = []
    for index, (header, body_blocks) in enumerate(sections):
        title = header.split(": ", 1)[1] if ": " in header else header
        body = "\n".join(body_blocks).strip()
        text = f"{title}\n{body}".strip() if body else title
        # 첫 헤딩 이전 preamble은 첫 섹션 도입부에 부착한다 (맥락 동봉).
        if index == 0 and preamble:
            text = "\n".join(preamble).strip() + "\n\n" + text
        drafts.append(ChunkDraft(text=text, section_header=header, is_atomic=False))
    return drafts


# --- 공통 엔트리 ---


def split_attachment(
    attachment: Attachment,
    attachment_type: AttachmentType | str,
) -> list[ChunkDraft]:
    """첨부 파일을 attachment_type별 전략으로 1차 분할한다.

    pdf는 폰트 휴리스틱 섹션, docx는 Heading 1/2/3 섹션, xlsx/csv는 N행 그룹을 1차 분할
    단위로 한다. 2차 크기 규칙은 chunk_attachment가 pdf/docx에 적용한다 (xlsx/csv는 행
    그룹 분할이 크기 처리를 겸한다).

    Args:
        attachment: 텍스트 추출 대상 첨부 객체. download_url이 실제 파일 경로를 가리킨다.
        attachment_type: 첨부 유형. pdf/docx/xlsx/csv를 지원한다.

    Returns:
        1차 분할 결과 ChunkDraft 목록. 본문이 비면 빈 목록.

    Raises:
        ValueError: 암호화된 PDF(ATTACH_ENCRYPTED) 등 처리할 수 없는 첨부.
    """
    resolved = _coerce_attachment_type(attachment_type)
    if resolved is AttachmentType.PDF:
        return _chunk_pdf(attachment)
    if resolved is AttachmentType.DOCX:
        return _chunk_docx(attachment)
    if resolved is AttachmentType.XLSX:
        return _chunk_xlsx(attachment)
    if resolved is AttachmentType.CSV:
        return _chunk_csv(attachment)
    raise ValueError(f"지원하지 않는 첨부 유형: {resolved}.")


def build_attachment_metadata(
    page: PageObject,
    attachment: Attachment,
    draft: ChunkDraft,
    chunk_index: int,
    attachment_type: AttachmentType,
) -> ChunkMetadata:
    """첨부 ChunkDraft에 청크 메타데이터 19종을 부착한다.

    무결성 규칙(chunking-strategy.md §6.3):
    - source_type=attachment, attachment_*·extracted_format 채움
    - doc_type 필드에는 첨부의 attachment_type 값을 담는다
    - chunk_id는 parent page_id + chunk_index + attachment_id 결정론적 SHA1
    - ACL·space_key·labels·webui_link는 부모 페이지에서 상속한다

    Args:
        page: 첨부의 부모 PageObject (ACL·메타 상속원).
        attachment: 청크의 출처 첨부 객체.
        draft: 1차 분할(필요 시 크기 규칙)을 거친 ChunkDraft.
        chunk_index: 동일 첨부 내 0-based 순번.
        attachment_type: 첨부 유형 (docx/xlsx).

    Returns:
        19종 필드가 채워진 ChunkMetadata.
    """
    section_header = draft.section_header.strip() or "untitled"
    section_path = " > ".join([*page.ancestors, attachment.filename, section_header])
    return ChunkMetadata(
        chunk_id=make_chunk_id(page.page_id, chunk_index, attachment.attachment_id),
        page_id=page.page_id,
        page_title=page.title,
        section_header=section_header,
        section_path=section_path,
        chunk_index=chunk_index,
        labels=page.labels,
        doc_type=attachment_type,
        space_key=page.space_key,
        allowed_groups=page.allowed_groups,
        allowed_users=page.allowed_users,
        webui_link=page.webui_link,
        last_modified=attachment.last_modified,
        source_type=SourceType.ATTACHMENT,
        attachment_id=attachment.attachment_id,
        attachment_filename=attachment.filename,
        attachment_mime=attachment.mime_type,
        extracted_format=_EXTRACTED_FORMAT_BY_TYPE[attachment_type],
        token_count=count_tokens(draft.text),
    )


def chunk_attachment(
    attachment: Attachment,
    page: PageObject,
    attachment_type: AttachmentType | str | None = None,
) -> list[Chunk]:
    """첨부 파일을 attachment_type별 전략으로 분할해 Chunk 목록을 반환한다.

    1차 분할 → (pdf/docx만)2차 크기 규칙 → 메타데이터 부착 순으로 처리한다.
    attachment_type이 None이면 mime/확장자 기반 PoC 휴리스틱(infer_attachment_type)으로
    추정한다.

    Args:
        attachment: 청킹 대상 첨부 객체. download_url이 실제 파일 경로를 가리킨다.
        page: 첨부의 부모 PageObject (ACL·메타 상속원).
        attachment_type: 첨부 유형. None이면 추정한다. pdf/docx/xlsx/csv를 지원한다.

    Returns:
        메타데이터가 부착된 Chunk 목록.

    Raises:
        ValueError: 암호화된 PDF(ATTACH_ENCRYPTED) 등 처리할 수 없는 첨부.
    """
    resolved = (
        _coerce_attachment_type(attachment_type)
        if attachment_type is not None
        else infer_attachment_type(attachment)
    )
    drafts = split_attachment(attachment, resolved)
    # pdf/docx 섹션은 비원자성이므로 2차 재분할·하한선 병합을 적용한다(헤딩 미검출 pdf의
    # 단일 draft는 여기서 800토큰 슬라이딩 윈도우로 재분할된다).
    # xlsx/csv는 행 그룹 분할이 크기 처리를 겸하므로 적용하지 않는다.
    if resolved in (AttachmentType.PDF, AttachmentType.DOCX):
        drafts = apply_size_rules(drafts)
    return [
        Chunk(
            text=draft.text,
            metadata=build_attachment_metadata(page, attachment, draft, index, resolved),
        )
        for index, draft in enumerate(drafts)
        if draft.text.strip()
    ]
